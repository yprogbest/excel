import openpyxl as op
from openpyxl.chart import LineChart, Reference, Series



if __name__ == "__main__":

    #グラフの作成
    #https://atmarkit.itmedia.co.jp/ait/articles/2203/08/news028.html
    #https://www.shibutan-bloomers.com/python_libraly_openpyxl-9/3126/

    # wb = op.load_workbook("/home/ubuntu/conda_src/excel/data/test_ver0.xlsx")# ワークシートの読み込み
    wb = op.load_workbook("/home/ubuntu/conda_src/excel/data/sample_chart.xlsx")# ワークシートの読み込み
    ws = wb['Sheet1'] # ワークシートの有効化
    # ws = wb['Sheet2'] # ワークシートの有効化


    graph_width = 12
    graph_height = 6


    # 1つ目のデータ
    #データ範囲の設定
    min_col_A = 2
    max_col_A = 7
    min_row_A = 2
    max_row_A = 5

    chart_A = LineChart()

    cats_A = Reference(ws, min_col=min_col_A+1, min_row=min_row_A, max_col=max_col_A, max_row=min_row_A)

    # 各行のデータを系列として追加
    for i in range(min_row_A+1, max_row_A+1):
        data_A = Reference(ws, min_col=min_col_A+1, min_row=i, max_col=max_col_A, max_row=i)
        series_A = Series(data_A, title=ws.cell(row=i, column=min_col_A).value)
        chart_A.append(series_A)

    chart_A.set_categories(cats_A)
    chart_A.anchor = 'I'+str(min_row_A)  # グラフの表示位置
    chart_A.width = graph_width  # グラフのサイズ
    chart_A.height = graph_height

    # 各シリーズに自動マーカーを設定
    for s_A in chart_A.series:
        s_A.marker.symbol = "auto"
        s_A.smooth = False

    ws.add_chart(chart_A)
    




    # 2つ目のデータ
    #データ範囲の設定
    min_col_B = 2
    max_col_B = 7
    min_row_B = 16
    max_row_B = 19

    chart_B = LineChart()

    cats_B = Reference(ws, min_col=min_col_B+1, min_row=min_row_B, max_col=max_col_B, max_row=min_row_B)

    # 各行のデータを系列として追加
    for i in range(min_row_B+1, max_row_B+1):
        data_B = Reference(ws, min_col=min_col_B+1, min_row=i, max_col=max_col_B, max_row=i)
        series_B = Series(data_B, title=ws.cell(row=i, column=min_col_B).value)
        chart_B.append(series_B)

    chart_B.set_categories(cats_B)
    chart_B.anchor = 'I'+str(min_row_B) # グラフの表示位置
    chart_B.width = graph_width # グラフのサイズ
    chart_B.height = graph_height

    # 各シリーズに自動マーカーを設定
    for s_B in chart_B.series:
        s_B.marker.symbol = "auto"
        s_B.smooth = False

    ws.add_chart(chart_B)






    wb.save("/home/ubuntu/conda_src/excel/data/sample_chart.xlsx")


