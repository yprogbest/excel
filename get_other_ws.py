import openpyxl
import glob
import os
from datetime import datetime


# 「レポート」シートを格納したExcelファイルを取得する
'''
param
- rootName: ルートディレクトリのパス
- id: 対象のID
- stDate: 開始年月日
return
- repfileList: 「レポート」シートを格納したExcelファイル
'''
def get_repfile(rootName, id, stDate):
    dname = f'{rootName}/{id}/{stDate}'
    repfileList = glob.glob(f'{dname}/rep_*.xlsx')
    return repfileList

#  「レポート」シートを格納したExcelファイルから、「レポート」シートを取得する
'''
param
- repfileList: 「レポート」シートを格納したExcelファイル
return
- repwsList: 「レポート」シートを格納したリスト
'''
def get_repws(repfileList):
    repwsList = []
    for repfile in repfileList:
        repwb = openpyxl.load_workbook(repfile)
        repws = repwb['レポート']
        repwsList.append(repws)
    return repwsList

# main.xlsxに新しくシートを作成し、「レポート」シートの内容をコピーする
'''
param
- mainFileName: main.xlsxファイルの絶対パス
- repwsList: 「レポート」シートを格納したリスト
'''
def write_repdata(mainFileName, repwsList):
    mainwb = openpyxl.load_workbook(mainFileName) # main.xlsxファイルをロードする

    for repws in repwsList:
        dtime_str = repws['A1'].value # 年月日時分秒
        dtime = datetime.strptime(dtime_str, '%Y-%m-%d %H:%M:%S')
        dyear = dtime.year # 年
        dmonth = dtime.month # 月
        dday = dtime.day # 日
        
        repwsName = f'{dyear:04}_{dmonth:02}_{dday:02}' # main.xlsxに新しく作成するシート名
        # 既にシートが存在するか確認する
        if repwsName not in mainwb.sheetnames:
            mainwb.create_sheet(title=repwsName) # main.xlsxに新しいシートを作成する
            # main.xlsxに作成したシートに「レポート」シートの内容をコピーする
            for rowNum, row in enumerate(repws.rows):
                for colNum, col in enumerate(row):
                    # main.xlsxに新しく作成したシートに、「レポート」シートの内容をコピーする
                    mainwb[repwsName].cell(row=rowNum+1, column=colNum+1).value = col.value
        else:
            print(f'{repwsName}は既に存在しています。')

        mainwb.save(mainFileName) # main.xlsxを保存する







if __name__ == '__main__':
    mainFileName = '/home/ubuntu/conda_src/excel/!他のExcelファイルからシートを取得/data/main.xlsx' # main.xlsxファイルの絶対パス
    
    rootName = os.path.dirname(mainFileName) # ルートディレクトリのパス
    id = '1234-5678'
    stDate = '20241007'

    repfileList = get_repfile(rootName, id, stDate) # 「レポート」シートを格納したExcelファイルのリスト
    repwsList = get_repws(repfileList) # 「レポート」シートのリスト
    write_repdata(mainFileName, repwsList) # main.xlsxに新しくシートを作成し、「レポート」シートの内容をコピーする