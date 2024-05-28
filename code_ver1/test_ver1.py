import openpyxl as op
from datetime import datetime



if __name__ == "__main__":


    wb = op.load_workbook("/home/ubuntu/conda_src/excel/data/test_ver1/test_ver1.xlsx", data_only=False)# ワークシートの読み込み
    wb_temp = op.load_workbook("/home/ubuntu/conda_src/excel/data/test_ver1/test_ver1.xlsx", data_only=True)# ワークシートの読み込み
    main_sheet_name = "Sheet1"
    ws = wb[main_sheet_name] # ワークシートの有効化
    ws_temp = wb_temp[main_sheet_name] # ワークシートの有効化
    

    # 各シートをリストに格納
    each_date_sheet_list = []
    for i in range(12, 17):
        sheet_name = f'2024_04_{i}' #シートの名前

        if sheet_name in wb.sheetnames:
            each_date_sheet = wb_temp[f'2024_04_{i}']
        else:
            each_date_sheet = "NoneSheet"

        each_date_sheet_list.append(each_date_sheet)






    #各シートの統計値をSheet1に入力する
    for j in range(len(each_date_sheet_list)):

        dic = {}
        col = 2

        # 各日のシートが存在するなら
        if each_date_sheet_list[j] != "NoneSheet":

            while(str(each_date_sheet_list[j].cell(column=col, row=10).value) != "None"): #列方向

                row_ = 11

                #リストの削除
                value_list = []  # 新しいリストを作成
                while(str(each_date_sheet_list[j].cell(column=col, row=row_).value) != "None"):

                    value_list.append(each_date_sheet_list[j].cell(column=col, row=row_).value) #値を取得


                    row_ +=1

                #辞書にキーと値を格納
                dic[each_date_sheet_list[j].cell(column=col, row=10).value] = value_list

                col += 1


            # print(f'dic = {dic}')
            # print(f'dic size = {len(dic)}')


            
            #辞書のキーを用いて、Sheet1のB列を探索
            for key, value in dic.items():

                flag = False
                for row in ws_temp.iter_rows(min_col=2, max_col=2, min_row=1, max_row=ws_temp.max_row): #B列の1行目から一番下までをfor文で走査する
                    for cell in row:

                        if cell.value == key: #もし、Sheet1のB列の値と辞書のキーが等しいなら

                            col_ = cell.column + (j+1)
                            row_ = cell.row + 2

                            #もし、日時の下のデータの値が0でないなら、
                            if ws.cell(column=col_, row=4).value != 0:
                                for i in range(len(value)): #統計値をSheet1に入力していく
                                    ws.cell(column=col_, row=row_).value = value[i]
                                    row_ += 1




                            # print(f'cell.column = {cell.column}') #セル番地の取得
                            # print(f'cell.row = {cell.row}') #セル番地の取得
                            # print(f'cell.value = {cell.value}') # セルの値
                            flag = True

                            break

                    if flag == True:
                        break





    wb.save("/home/ubuntu/conda_src/excel/data/test_ver1/output.xlsx")

    print("書き込みが完了しました。")